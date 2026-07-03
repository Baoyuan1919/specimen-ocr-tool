#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
标本照片 → Excel 自动填表 v2.1
- 字段自定义增删（设置中自由增减，也可从 Excel 表头同步）
- 系统托盘后台运行（关闭窗口最小化到托盘区）
- 文件夹自动监听（检测新图片自动识别填表）
- 基于坐标的字段匹配（完美适配表格排版照片）
"""

import os, re, sys, json, threading, time, csv
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

try:
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox, scrolledtext
except ImportError:
    tk = None

# 默认字段
DEFAULT_FIELDS = [
    "采集号", "采集时间", "采集人", "采集地点",
    "经度", "纬度", "海拔",
    "习性", "生态环境", "高度"
]

SUPPORTED_EXT = {'.jpg', '.jpeg', '.png', '.bmp', '.tiff', '.tif', '.webp'}
CONFIG_DIR = Path(os.environ.get("LOCALAPPDATA", Path.home())) / "标本OCR工具"
CONFIG_FILE = CONFIG_DIR / "config.json"
LOG_FILE = CONFIG_DIR / "操作日志.csv"


# ─── OCR 引擎 ───
_ocr = None
def get_ocr():
    global _ocr
    if _ocr is None:
        from rapidocr_onnxruntime import RapidOCR
        _ocr = RapidOCR()
    return _ocr

def is_img(path):
    return Path(path).suffix.lower() in SUPPORTED_EXT


def ocr_image(img_path):
    """
    OCR识别，返回 (items, raw_texts)
    items: [(text, x_center, y_center, box_width), ...] 按位置排序
    raw_texts: [text, ...] 纯文本列表（旧版兼容）
    """
    ocr = get_ocr()
    result, _ = ocr(str(img_path))
    items = []
    raw_texts = []
    if result:
        for box in result:
            txt = box[1].strip()
            if not txt:
                continue
            raw_texts.append(txt)
            coords = box[0]  # [[x1,y1],[x2,y2],[x3,y3],[x4,y4]]
            xs = [p[0] for p in coords]
            ys = [p[1] for p in coords]
            x_center = sum(xs) / 4
            y_center = sum(ys) / 4
            box_w = max(xs) - min(xs)
            items.append((txt, x_center, y_center, box_w))
    # 按垂直位置分组排序，同行的按水平排序
    # 用 y 坐标的 20px 作为"同行的容差"
    items.sort(key=lambda it: (round(it[2] / 20) * 20, it[1]))
    return items, raw_texts


# 需要严格精确匹配的单字字段（避免误匹配内容中的同字）
_STRICT_FIELDS = {'茎', '叶', '花', '果', '根', '皮', '枝', '芽'}


def _is_field_match(field, txt):
    """判断 OCR 文本 txt 是否匹配用户设置的字段名 field。

    规则：
    1. 精确匹配总是优先
    2. 单字字段（茎/叶/花/果等）：
       - 如果 OCR 文本含括号（如「枝（茎）」），检查括号内是否含字段名
       - 不允许 OCR 文本以字段名开头但却是长内容（如「茎生叶」不是 field「茎」）
    3. 多字字段：如果短 OCR 文本（≤3字符）是字段名的子串则匹配（如「果」→「果实」）
    4. 其他字段（科/属/采集号等）：支持前缀匹配（科→科名）
    """
    t = txt.strip('：: ')
    if not t:
        return False
    # 如果有冒号，取冒号前的内容
    txt_before_colon = txt.split('：')[0].split(':')[0].strip() if '：' in txt or ':' in txt else None
    compare = txt_before_colon if txt_before_colon else t

    # 规则1：精确匹配
    if compare == field:
        return True

    # 规则2：单字字段特殊处理
    if field in _STRICT_FIELDS:
        # 含括号的短文本（如「枝（茎）」），查括号内是否含字段名
        if '（' in txt and '）' in txt:
            paren_content = txt[txt.find('（') + 1:txt.find('）')]
            if field in paren_content:
                return True
        # 字段名出现在文本末尾（如「花果」→字段「果」）但没有内容描述特征
        if 2 <= len(compare) <= 4:
            # 短文本且字段在末尾
            if compare.endswith(field) and not any(k in compare for k in ['生', '形', '状', '色', '毛']):
                return True
        return False

    if len(field) >= 2 and len(compare) <= 3 and len(compare) >= 1:
        # 规则3：短 OCR 文本是字段名的子串（「果」→「果实」）
        if compare in field or field in compare:
            return True

    # 规则4：其他字段前缀匹配
    if txt_before_colon:
        return txt_before_colon.startswith(field) or txt_before_colon == field
    return t.startswith(field) or t == field


def parse_fields(ocr_items, raw_texts, fields):
    """
    基于坐标位置的字段匹配，适配表格和标签两种排版。

    匹配策略（优先级从高到低）：
    1. 同文本冒号提取：字段名本身包含冒号（如「采集号：629022105203」）
    2. 坐标邻近匹配：字段名右侧且同一水平区域的内容
    3. 全文本正则回退

    特点：
    - 单字字段（茎/叶/花/果）仅精确匹配，避免误取内容
    - 多字字段（科→科名）支持前缀匹配
    - 重复字段仅匹配一次
    - 值中混入其他字段名时自动截断
    """
    result = {}
    used_positions = {}  # 已匹配的字段名位置，避免重复
    field_match_log = {}  # 调试：记录匹配了什么

    for field in fields:
        val = ""

        # 找出字段名在图片中的所有出现位置
        field_matches = []
        for txt, x, y, bw in ocr_items:
            if _is_field_match(field, txt):
                field_matches.append((txt, x, y, bw))

        if not field_matches:
            # 字段名在 OCR 结果中完全找不到 → 回退到正则
            full = '\n'.join(raw_texts)
            # 支持「科名」匹配字段「科」这种场景
            m = re.search(r'(?:' + re.escape(field) + r'|' +
                          re.escape(field) + r'名)' +
                          r'[\s]*[:：]\s*([^\n]{1,200})', full)
            if not m:
                m = re.search(re.escape(field) + r'[\s]*[:：]\s*([^\n]{1,200})', full)
            if m:
                v = m.group(1).strip().rstrip('，。.;,;）)')
                if v:
                    val = v
            result[field] = val
            continue

        # ── 策略1：看字段名本身是否包含冒号+值 ──
        for txt, x, y, bw in field_matches:
            if '：' in txt:
                parts = txt.split('：', 1)
                if len(parts) > 1 and parts[1].strip():
                    val = parts[1].strip().rstrip('，。.;,;）)')
                    used_positions[field] = (x, y)
                    field_match_log[field] = f"冒号提取: {txt}"
                    break
            elif ':' in txt:
                parts = txt.split(':', 1)
                if len(parts) > 1 and parts[1].strip():
                    val = parts[1].strip().rstrip('，。.;,;）)')
                    used_positions[field] = (x, y)
                    field_match_log[field] = f"冒号提取: {txt}"
                    break

        # ── 策略2：字段名右侧或下方找值（表格+标签排版） ──
        if not val:
            for txt, fx, fy, fbw in field_matches:
                if field in used_positions:
                    continue
                right_edge = fx + fbw
                candidates = []
                for vt, vx, vy, vbw in ocr_items:
                    # 跳过字段名自己
                    if _is_field_match(field, vt) and abs(vy - fy) < 15:
                        continue
                    # 跳过已匹配的其他字段
                    skip = False
                    for ff, (ffx, ffy) in used_positions.items():
                        if _is_field_match(ff, vt) and abs(vy - ffy) < 20:
                            skip = True
                            break
                    if skip:
                        continue

                    # A) 右侧匹配（同行水平带，字段名右侧 0~600px）
                    if abs(vy - fy) <= 50:
                        dx = vx - right_edge
                        if -5 < dx < 600:
                            candidates.append((1, dx, vt))
                    # B) 下方匹配（下一行，垂直 20~100px，水平偏移 ±150px）
                    dy = vy - fy
                    if 20 < dy < 100 and abs(vx - fx) < 150:
                        dx = vx - fx
                        if dx > -20:
                            candidates.append((2, dy, vt))

                if candidates:
                    candidates.sort(key=lambda c: (c[0], c[1]))
                    raw_val = candidates[0][2].strip()
                    raw_val = raw_val.split('\n')[0].strip()
                    raw_val = raw_val.rstrip('，。.;,;）)').rstrip('：:')
                    # 如果值中还包含其他字段名，截断
                    for f2 in fields:
                        if f2 == field:
                            continue
                        # 单字字段要小心（值里可能正好包含这个字）
                        if len(field) == 1:
                            continue
                        idx = raw_val.find(f2)
                        if idx > 3:
                            raw_val = raw_val[:idx].strip()
                            break
                    if raw_val:
                        val = raw_val
                        used_positions[field] = (fx, fy)
                        break

        # ── 策略3：全文本正则回退（支持科名→科） ──
        if not val:
            full = '\n'.join(raw_texts)
            m = re.search(r'(?:' + re.escape(field) + r'|' +
                          re.escape(field) + r'名)' +
                          r'[\s]*[:：]\s*([^\n]{1,200})', full)
            if not m:
                m = re.search(re.escape(field) + r'[\s]*[:：]\s*([^\n]{1,200})', full)
            if m:
                v = m.group(1).strip().rstrip('，。.;,;）)')
                if v:
                    val = v

        result[field] = val

    return result


# ─── Excel 操作 ───

def create_excel(path, fields):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "标本数据"
    hf = Font(bold=True, size=11, color="FFFFFF")
    hfill = PatternFill("solid", fgColor="4472C4")
    halign = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bdr = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    headers = ["序号"] + fields
    widths = [8] + [16] * len(fields)
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font, c.fill, c.alignment, c.border = hf, hfill, halign, bdr
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'
    wb.save(path)
    return path


def write_row(excel_path, row_data, field_order, row_num):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    bdr = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    c = ws.cell(row=row_num, column=1, value=row_num - 1); c.border = bdr
    for i, field in enumerate(field_order, 2):
        v = row_data.get(field, "") or ""
        c = ws.cell(row=row_num, column=i, value=v); c.border = bdr
    wb.save(excel_path)


def write_operation_log(excel_path, total, succeed, rename_count, fields_count):
    """写入操作日志，每天一条累计记录"""
    try:
        today = date.today().isoformat()
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        rows = []
        # 读取已有日志
        if LOG_FILE.exists():
            try:
                with open(LOG_FILE, 'r', encoding='utf-8', newline='') as f:
                    reader = csv.DictReader(f)
                    for r in reader:
                        rows.append(r)
            except Exception:
                pass

        # 找今天的记录
        today_row = None
        for r in rows:
            if r.get("日期") == today:
                today_row = r
                break

        if today_row:
            # 更新今天的记录
            prev_total = int(today_row.get("累计标本数", 0))
            # 重新统计 Excel 中的实际行数
            actual = count_data_rows(excel_path) if os.path.isfile(excel_path) else succeed
            today_row["累计标本数"] = str(actual)
            today_row["最后操作"] = now
            # 更新当日新增（如果 Excel 统计 > 之前累计，差额即为新增）
            prev_day_total = int(today_row.get("当日新增", 0))
            today_row["当日新增"] = str(prev_day_total + succeed)
            today_row["成功数"] = str(int(today_row.get("成功数", 0)) + succeed)
            today_row["失败数"] = str(int(today_row.get("失败数", 0)) + (total - succeed))
            log_fn = today_row.get("使用的Excel", "")
            if excel_path and excel_path not in log_fn:
                log_fn = log_fn + "; " + excel_path if log_fn else excel_path
            today_row["使用的Excel"] = log_fn
        else:
            # 新建今天的记录
            actual = succeed
            if os.path.isfile(excel_path):
                actual = count_data_rows(excel_path)
            rows.append({
                "日期": today,
                "累计标本数": str(actual),
                "当日新增": str(succeed),
                "成功数": str(succeed),
                "失败数": str(total - succeed),
                "重命名数": str(rename_count),
                "字段数": str(fields_count),
                "使用的Excel": excel_path or "",
                "最后操作": now
            })

        # 写回
        headers = ["日期","累计标本数","当日新增","成功数","失败数","重命名数","字段数","使用的Excel","最后操作"]
        with open(LOG_FILE, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            writer.writerows(rows)
    except Exception:
        pass


def count_data_rows(excel_path):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    return max(0, ws.max_row - 1)


def get_excel_headers(excel_path):
    """读取 Excel 表头字段（跳过「序号」列）"""
    if not os.path.isfile(excel_path):
        return None
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        headers = []
        for c in range(2, ws.max_column + 1):
            v = ws.cell(row=1, column=c).value
            if v and str(v).strip() and str(v).strip() != "序号":
                headers.append(str(v).strip())
        return headers if headers else None
    except Exception:
        return None


# ─── 主应用 ───
class App:
    def __init__(self):
        if tk is None:
            import ctypes
            ctypes.windll.user32.MessageBoxW(0, "需要 tkinter（Windows 自带 Python 已包含）", "错误", 0)
            sys.exit(1)

        self.root = tk.Tk()
        self.root.title("标本照片 → Excel 自动填表")
        self.root.withdraw()

        self.fields = DEFAULT_FIELDS.copy()
        self.image_folder = tk.StringVar()
        self.excel_path = tk.StringVar()
        self.running = False
        self.monitoring = False
        self.monitor_thread = None
        self._stop_monitor = threading.Event()

        self._load_config()
        self._build_ui()

        self.root.deiconify()
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)
        self.root.mainloop()

    def _build_ui(self):
        self.root.geometry("780x680")
        self.root.minsize(640, 520)

        m = ttk.Frame(self.root, padding=12)
        m.pack(fill=tk.BOTH, expand=True)

        ttk.Label(m, text="标本照片 OCR → Excel 自动填表",
                  font=('微软雅黑', 14, 'bold')).pack(pady=(0, 4))

        # ── 字段管理 ──
        ft = ttk.Frame(m)
        ft.pack(fill=tk.X, pady=(0, 4))
        self.fields_btn = ttk.Button(
            ft, text=f"⚙️ 管理字段（当前 {len(self.fields)} 个）",
            command=self._manage_fields, width=28)
        self.fields_btn.pack(side=tk.LEFT)
        ttk.Label(ft, text="增删改字段；可从 Excel 表头同步",
                  foreground="gray").pack(side=tk.LEFT, padx=8)

        # ── 重命名开关 ──
        f_rename = ttk.Frame(m)
        f_rename.pack(fill=tk.X, pady=(0, 4))
        self.rename_var = tk.BooleanVar(value=True)
        self.rename_cb = ttk.Checkbutton(
            f_rename, text="按采集号重命名图片",
            variable=self.rename_var,
            onvalue=True, offvalue=False)
        self.rename_cb.pack(side=tk.LEFT)
        ttk.Label(f_rename,
                  text="处理后将图片重命名为「采集号+.jpg」格式",
                  foreground="gray").pack(side=tk.LEFT, padx=8)

        # ── 图片文件夹 ──
        f1 = ttk.LabelFrame(m, text="📁 图片文件夹（按文件名排序处理）", padding=8)
        f1.pack(fill=tk.X, pady=3)
        r1 = ttk.Frame(f1)
        r1.pack(fill=tk.X)
        ttk.Entry(r1, textvariable=self.image_folder, width=55).pack(
            side=tk.LEFT, padx=(0, 6))
        ttk.Button(r1, text="浏览...", command=self._pick_folder,
                   width=8).pack(side=tk.LEFT)
        self.fs = ttk.Label(f1, foreground="gray")
        self.fs.pack(anchor=tk.W, pady=(3, 0))

        # ── Excel 输出 ──
        f2 = ttk.LabelFrame(m, text="📊 Excel 输出文件", padding=8)
        f2.pack(fill=tk.X, pady=3)
        r2 = ttk.Frame(f2)
        r2.pack(fill=tk.X)
        ttk.Entry(r2, textvariable=self.excel_path, width=55).pack(
            side=tk.LEFT, padx=(0, 6))
        ttk.Button(r2, text="选择已有...", command=self._pick_excel,
                   width=10).pack(side=tk.LEFT)
        ttk.Button(r2, text="新建...", command=self._new_excel,
                   width=8).pack(side=tk.LEFT, padx=4)
        self.es = ttk.Label(f2, foreground="gray")
        self.es.pack(anchor=tk.W, pady=(3, 0))

        self.image_folder.trace_add("write", lambda *a: self._refresh())
        self.excel_path.trace_add("write", lambda *a: self._refresh())
        self._refresh()

        # ── 日志 ──
        f3 = ttk.LabelFrame(m, text="📝 运行日志", padding=6)
        f3.pack(fill=tk.BOTH, expand=True, pady=3)
        self.log = scrolledtext.ScrolledText(
            f3, height=12, font=('Consolas', 10),
            bg="#1e1e1e", fg="#d4d4d4", insertbackground="white",
            state=tk.DISABLED)
        self.log.pack(fill=tk.BOTH, expand=True)
        for tag, color in [("i", "#d4d4d4"), ("ok", "#4ec9b0"),
                           ("err", "#f44747"), ("w", "#ce9178"),
                           ("b", "#569cd6")]:
            self.log.tag_config(tag, foreground=color)
        self.log.tag_config("b", font=('Consolas', 11, 'bold'))

        self.pb = ttk.Progressbar(m, mode='determinate')
        self.pb.pack(fill=tk.X, pady=2)
        self.pl = ttk.Label(m, text="", foreground="gray")
        self.pl.pack()

        # ── 按钮区 ──
        bf = ttk.Frame(m)
        bf.pack(fill=tk.X, pady=(4, 0))

        self.run_btn = ttk.Button(
            bf, text="🚀 手动识别并填表", width=20, command=self._start)
        self.run_btn.pack(side=tk.RIGHT, padx=3)

        self.monitor_btn = ttk.Button(
            bf, text="👁 开始监听", width=14, command=self._toggle_monitor)
        self.monitor_btn.pack(side=tk.RIGHT, padx=3)

        ttk.Button(bf, text="清空日志", width=10,
                   command=lambda: (self.log.config(state=tk.NORMAL),
                                    self.log.delete('1.0', tk.END),
                                    self.log.config(state=tk.DISABLED))
                   ).pack(side=tk.RIGHT, padx=3)

        ttk.Button(bf, text="打开输出目录", width=12,
                   command=self._open_dir).pack(side=tk.LEFT, padx=3)

        self.tray_btn = ttk.Button(
            bf, text="— 最小化到托盘", width=14,
            command=self._minimize_to_tray)
        self.tray_btn.pack(side=tk.LEFT, padx=3)

        ttk.Button(bf, text="📊 统计", width=8,
                   command=self._show_stats).pack(side=tk.LEFT, padx=3)

    # ─── 日志 / 刷新 ───
    def _wl(self, msg, tag="i"):
        self.log.config(state=tk.NORMAL)
        self.log.insert(tk.END, msg + "\n", tag)
        self.log.see(tk.END)
        self.log.config(state=tk.DISABLED)
        self.root.update_idletasks()

    def _refresh(self):
        d, e = self.image_folder.get(), self.excel_path.get()
        n = len([p for p in Path(d).iterdir() if is_img(p)]) \
            if d and os.path.isdir(d) else 0
        self.fs.config(
            text=f"已选（{n} 张图片）" if n else "（未选择）",
            foreground="green" if n else "gray")
        self.es.config(
            text="✅ 文件存在" if e and os.path.isfile(e)
            else ("⚠️ 不存在，自动创建" if e else "（未选择）"),
            foreground="green" if e and os.path.isfile(e)
            else ("orange" if e else "gray"))

    # ─── 路径选择 ───
    def _pick_folder(self):
        d = filedialog.askdirectory()
        d and self.image_folder.set(d)

    def _pick_excel(self):
        p = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        p and self.excel_path.set(p)

    def _new_excel(self):
        p = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
        if p:
            create_excel(p, self.fields)
            self.excel_path.set(p)
            self._wl(f"📋 已创建: {Path(p).name}", "ok")

    def _open_dir(self):
        p = self.excel_path.get()
        p and os.path.isfile(p) and os.startfile(os.path.dirname(p))

    # ─── 字段管理 ───
    def _manage_fields(self):
        win = tk.Toplevel(self.root)
        win.title("字段管理")
        win.geometry("440x520")
        win.transient(self.root)
        win.grab_set()

        ttk.Label(win, text="管理识别字段",
                  font=('微软雅黑', 12, 'bold')).pack(pady=(10, 2))
        ttk.Label(
            win,
            text="每行一个字段名，需与照片上的文字名称完全一致",
            foreground="gray").pack()
        ttk.Label(
            win,
            text='例如照片上写「采集号」→ 字段名就写「采集号」',
            foreground="gray").pack(pady=(0, 6))

        text = scrolledtext.ScrolledText(win, height=14, font=('微软雅黑', 10))
        text.pack(fill=tk.BOTH, expand=True, padx=12, pady=4)
        for f in self.fields:
            text.insert(tk.END, f + "\n")

        def sync_from_excel():
            e = self.excel_path.get()
            if not e or not os.path.isfile(e):
                messagebox.showinfo("提示", "请先选择 Excel 文件", parent=win)
                return
            headers = get_excel_headers(e)
            if not headers:
                messagebox.showinfo("提示", "Excel 中未找到有效表头", parent=win)
                return
            text.delete('1.0', tk.END)
            for h in headers:
                text.insert(tk.END, h + "\n")
            self._wl(f"📋 已从 Excel 同步 {len(headers)} 个字段", "ok")

        def save_fields():
            raw = text.get('1.0', tk.END).strip()
            new_fields = [l.strip() for l in raw.split('\n') if l.strip()]
            if not new_fields:
                messagebox.showwarning("提示", "至少保留一个字段", parent=win)
                return
            self.fields = new_fields
            self._save_config()
            self.fields_btn.config(
                text=f"⚙️ 管理字段（当前 {len(self.fields)} 个）")
            win.destroy()
            n = len(self.fields)
            disp = ', '.join(self.fields[:5])
            self._wl(
                f"✅ 字段已更新（{n} 个）：{disp}{'...' if n > 5 else ''}",
                "ok")

        bf = ttk.Frame(win)
        bf.pack(fill=tk.X, padx=12, pady=6)

        ttk.Button(bf, text="从 Excel 表头同步",
                   command=sync_from_excel).pack(side=tk.LEFT)
        ttk.Button(bf, text="重置默认",
                   command=lambda: (text.delete('1.0', tk.END),
                                    text.insert(
                                        tk.END,
                                        "\n".join(DEFAULT_FIELDS) + "\n"))
                   ).pack(side=tk.LEFT, padx=6)

        ttk.Button(bf, text="✅ 保存", width=10,
                   command=save_fields).pack(side=tk.RIGHT, padx=4)
        ttk.Button(bf, text="取消", width=8,
                   command=win.destroy).pack(side=tk.RIGHT, padx=4)

    # ─── 系统托盘 ───
    def _minimize_to_tray(self):
        try:
            from PIL import Image, ImageDraw, ImageFont
            import pystray

            self.root.withdraw()
            self._wl("🔽 已最小化到系统托盘，双击图标恢复窗口", "ok")

            img = Image.new('RGBA', (64, 64), (0, 0, 0, 0))
            draw = ImageDraw.Draw(img)
            draw.ellipse([2, 2, 62, 62], fill=(68, 114, 196))
            try:
                font = ImageFont.truetype("msyh.ttc", 28)
            except Exception:
                font = ImageFont.load_default()
            draw.text((16, 16), "标", fill=(255, 255, 255), font=font)

            def on_show(icon, item):
                icon.stop()
                self.root.after(0, self.root.deiconify)

            def on_run(icon, item):
                icon.stop()
                self.root.after(
                    0, lambda: (self.root.deiconify(), self._start()))

            def on_quit(icon, item):
                icon.stop()
                self._save_config()
                if self.monitoring:
                    self._stop_monitor.set()
                self.root.after(0, self._quit)

            menu = pystray.Menu(
                pystray.MenuItem("显示窗口", on_show, default=True),
                pystray.MenuItem("立即识别", on_run),
                pystray.Menu.SEPARATOR,
                pystray.MenuItem("退出", on_quit),
            )
            icon = pystray.Icon("specimen_ocr", img, "标本OCR填表工具", menu)
            threading.Thread(target=icon.run, daemon=True).start()

        except ImportError:
            self.root.iconify()
            self._wl("📌 已最小化到任务栏"
                      "（安装 pystray 可支持系统托盘）", "i")
        except Exception as e:
            self.root.iconify()
            self._wl(f"📌 已最小化到任务栏（托盘异常: {e}）", "w")

    # ─── 文件夹监听 ───
    def _toggle_monitor(self):
        if self.monitoring:
            self._stop_monitor.set()
            self.monitoring = False
            self.monitor_btn.config(text="👁 开始监听")
            self._wl("🛑 已停止文件夹监听", "w")
            return

        d = self.image_folder.get()
        if not d or not os.path.isdir(d):
            messagebox.showwarning("提示", "请先选择图片文件夹")
            return

        self._stop_monitor.clear()
        self.monitoring = True
        self.monitor_btn.config(text="⏹ 停止监听")
        self._wl("👁 开始监听文件夹", "ok")
        self._wl(f"   📂 {d}", "i")
        self._wl("   → 检测到新图片将自动识别并填入 Excel", "i")

        def monitor_loop():
            processed = set()
            try:
                for p in Path(d).iterdir():
                    if is_img(p):
                        processed.add(p.name)
            except Exception:
                pass
            self.root.after(
                0, lambda: self._wl(
                    f"   ℹ️ 已有 {len(processed)} 张图片，仅处理新增的", "i"))

            while not self._stop_monitor.is_set():
                time.sleep(3)
                try:
                    e = self.excel_path.get()
                    if not e or not os.path.isdir(d):
                        continue
                    current = sorted(
                        [p for p in Path(d).iterdir() if is_img(p)])
                    new_imgs = [p for p in current
                                if p.name not in processed]
                    if not new_imgs:
                        continue

                    fields = self.fields
                    hdrs = get_excel_headers(e)
                    if hdrs:
                        fields = hdrs

                    for img in new_imgs:
                        if self._stop_monitor.is_set():
                            break
                        name = img.name
                        self.root.after(
                            0, lambda n=name: self._wl(
                                f"[监听] 📷 {n}", "ok"))
                        try:
                            items, raw = ocr_image(str(img))
                            if items:
                                row_data = parse_fields(items, raw, fields)
                                base = count_data_rows(e)
                                if not os.path.isfile(e):
                                    create_excel(e, fields)
                                write_row(e, row_data, fields, base + 2)
                                processed.add(name)
                                self.root.after(
                                    0, lambda n=name: self._wl(
                                        f"[监听] ✅ {n} → 已填入 Excel",
                                        "ok"))
                            else:
                                processed.add(name)
                                self.root.after(
                                    0, lambda n=name: self._wl(
                                        f"[监听] ⚠️ {n} 未识别到文字",
                                        "w"))
                        except Exception as ex:
                            self.root.after(
                                0, lambda n=name, e=str(ex): self._wl(
                                    f"[监听] ❌ {n}: {e}", "err"))
                except Exception:
                    pass

        self.monitor_thread = threading.Thread(
            target=monitor_loop, daemon=True)
        self.monitor_thread.start()

    # ─── 手动识别 ───
    def _start(self):
        if self.running:
            return
        d, e = self.image_folder.get(), self.excel_path.get()
        if not d or not os.path.isdir(d):
            messagebox.showwarning("提示", "请选择图片文件夹")
            return
        if not e:
            messagebox.showwarning("提示", "请选择或新建 Excel")
            return

        self.running = True
        self.run_btn.config(state=tk.DISABLED, text="⏳ 处理中...")
        self.pb['value'] = 0
        self.pl.config(text="正在初始化 OCR 引擎...")
        self._wl("=" * 45, "b")
        self._wl("🚀 开始处理标本照片...", "b")
        self._wl("=" * 45, "b")

        def worker():
            fields = self.fields
            hdrs = get_excel_headers(e)
            if hdrs:
                fields = hdrs
                if hdrs != self.fields:
                    self.root.after(
                        0, lambda: self._wl(
                            f"   ℹ️ 字段已与 Excel 表头同步"
                            f"（{len(hdrs)} 个）", "i"))

            def pc(c, t):
                self.root.after(0, lambda: (
                    self.pb.configure(value=c, maximum=t),
                    self.pl.config(
                        text=f"处理中 {c}/{t} [{c*100//t}%]")
                ))

            def lc(m, tag="i"):
                self.root.after(0, lambda: self._wl(m, tag))

            try:
                images = sorted(
                    [p for p in Path(d).iterdir() if is_img(p)])
                total = len(images)
                if total == 0:
                    lc("⚠️ 未找到图片文件", "w")
                    return
                if not os.path.isfile(e):
                    create_excel(e, fields)
                    lc(f"📋 已创建 Excel 模板: {Path(e).name}", "ok")

                base_rows = count_data_rows(e)
                succeed = 0
                rename_count = 0

                for idx, img in enumerate(images):
                    name = img.name
                    lc(f"[{idx+1}/{total}] 📷 {name}")
                    pc(idx + 1, total)
                    try:
                        items, raw = ocr_image(str(img))
                        if not items:
                            lc(f"   ⚠️ 未识别到文字", "w")
                            continue
                        row_data = parse_fields(items, raw, fields)
                        for f in fields:
                            v = row_data.get(f, "") or "（未识别）"
                            lc(f"   · {f} → {v}")
                        row_num = base_rows + succeed + 2
                        write_row(e, row_data, fields, row_num)
                        succeed += 1
                        lc(f"   ✅ 已写入 → 行{row_num}", "ok")
                        # 按采集号重命名图片
                        if self.rename_var.get():
                            caihao = row_data.get("采集号", "") or ""
                            if caihao and caihao != "（未识别）":
                                safe = re.sub(r'[\\/:*?"<>|]', '_', caihao).strip()
                                if safe:
                                    ext = img.suffix.lower()
                                    # 所有图片均从 _01 开始编号
                                    counter = 1
                                    while True:
                                        new_name = f"{safe}_{counter:02d}{ext}"
                                        new_path = img.parent / new_name
                                        if not new_path.exists():
                                            break
                                        counter += 1
                                    img.rename(new_path)
                                    rename_count += 1
                                    lc(f"   📎 已重命名为 {new_name}", "ok")
                    except Exception as ex:
                        lc(f"   ❌ 处理失败: {ex}", "err")

                lc(f"\n{'=' * 45}", "b")
                lc(f"✅ 完成！共 {total} 张图片，"
                   f"成功写入 {succeed} 行", "ok")
                # 记录操作日志
                write_operation_log(e, total, succeed, rename_count, len(fields))
            except Exception as ex:
                lc(f"❌ 运行出错: {ex}", "err")
            finally:
                self.running = False
                self.root.after(0, self._done)

        threading.Thread(target=worker, daemon=True).start()

    def _done(self):
        self.run_btn.config(state=tk.NORMAL, text="🚀 手动识别并填表")
        self.pl.config(text="✅ 完成")
        self._save_config()

    def _show_stats(self):
        """显示统计面板"""
        win = tk.Toplevel(self.root)
        win.title("📊 录入统计")
        win.geometry("600x450")
        win.transient(self.root)

        ttk.Label(win, text="标本录入统计",
                  font=('微软雅黑', 14, 'bold')).pack(pady=(10, 4))

        # 从 Excel 实时统计
        e = self.excel_path.get()
        excel_count = 0
        if e and os.path.isfile(e):
            excel_count = count_data_rows(e)

        # 从日志读取历史
        log_rows = []
        if LOG_FILE.exists():
            try:
                with open(LOG_FILE, 'r', encoding='utf-8', newline='') as f:
                    reader = csv.DictReader(f)
                    for r in reader:
                        log_rows.append(r)
            except Exception:
                pass

        # 顶部：总体概览
        summary = ttk.LabelFrame(win, text="总体概览", padding=8)
        summary.pack(fill=tk.X, padx=12, pady=4)

        # Excel 实际行数
        ttk.Label(summary, text=f"当前 Excel 标本数：{excel_count}",
                  font=('微软雅黑', 11)).pack(anchor=tk.W)

        if log_rows:
            total_all = sum(int(r.get("累计标本数", 0)) for r in log_rows)
            total_ok = sum(int(r.get("成功数", 0)) for r in log_rows)
            total_fail = sum(int(r.get("失败数", 0)) for r in log_rows)
            total_rename = sum(int(r.get("重命名数", 0)) for r in log_rows)
            ttk.Label(summary,
                      text=f"操作总成功：{total_ok} 次  失败：{total_fail} 次"
                      ).pack(anchor=tk.W)
            ttk.Label(summary,
                      text=f"累计重命名：{total_rename} 张"
                      ).pack(anchor=tk.W)
            ttk.Label(summary,
                      text=f"日志天数：{len(log_rows)} 天"
                      ).pack(anchor=tk.W)

        # 日志列表
        f_log = ttk.LabelFrame(win, text="操作日志", padding=6)
        f_log.pack(fill=tk.BOTH, expand=True, padx=12, pady=4)

        columns = ("日期", "累计标本数", "当日新增", "成功", "失败", "重命名", "最后操作")
        tree = ttk.Treeview(f_log, columns=columns, show="headings", height=8)
        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=80 if col in ("日期","累计标本数","当日新增","成功","失败","重命名") else 160)
        tree.pack(fill=tk.BOTH, expand=True)

        # 从新到旧显示
        for r in reversed(log_rows):
            tree.insert("", "end", values=(
                r.get("日期",""),
                r.get("累计标本数","0"),
                r.get("当日新增","0"),
                r.get("成功数","0"),
                r.get("失败数","0"),
                r.get("重命名数","0"),
                r.get("最后操作","")
            ))

        if not log_rows:
            ttk.Label(f_log, text="暂无操作日志", foreground="gray").pack(pady=20)

        ttk.Button(win, text="关闭", width=10,
                   command=win.destroy).pack(pady=6)

    def _quit(self):
        self._save_config()
        if self.monitoring:
            self._stop_monitor.set()
        self.root.quit()
        self.root.destroy()

    def _on_close(self):
        self._save_config()
        self._minimize_to_tray()

    # ─── 配置持久化 ───
    def _save_config(self):
        try:
            CONFIG_DIR.mkdir(parents=True, exist_ok=True)
            Path(CONFIG_FILE).write_text(
                json.dumps({
                    "image_folder": self.image_folder.get(),
                    "excel_path": self.excel_path.get(),
                    "fields": self.fields
                }, ensure_ascii=False, indent=2),
                encoding='utf-8')
        except Exception:
            pass

    def _load_config(self):
        try:
            if CONFIG_FILE.exists():
                d = json.loads(
                    CONFIG_FILE.read_text(encoding='utf-8'))
                if isinstance(d.get("fields"), list) \
                        and len(d["fields"]) > 0:
                    self.fields = d["fields"]
                p1 = d.get("image_folder", "")
                if p1 and os.path.isdir(p1):
                    self.image_folder.set(p1)
                p2 = d.get("excel_path", "")
                if p2:
                    self.excel_path.set(p2)
        except Exception:
            pass


if __name__ == "__main__":
    try:
        from rapidocr_onnxruntime import RapidOCR
    except ImportError:
        print("=" * 50)
        print("📦 请先安装依赖：")
        print("   pip install rapidocr-onnxruntime openpyxl Pillow pystray")
        print("=" * 50)
        sys.exit(1)
    App()
